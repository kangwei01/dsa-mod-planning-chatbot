"""Streamlit user interface for the chatbot demo."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import requests
import streamlit as st
from openpyxl import Workbook, load_workbook

if __package__ in (None, ""):
    sys.path.append(str(Path(__file__).resolve().parent.parent))
    from app.chat_graph import DEFAULT_SYSTEM_PROMPT_TEMPLATE
else:
    from .chat_graph import DEFAULT_SYSTEM_PROMPT_TEMPLATE

API_ROOT = os.getenv("CHATBOT_API_ROOT", "http://localhost:5000/api")
CHAT_ENDPOINT = f"{API_ROOT}/chat"
RESET_ENDPOINT = f"{API_ROOT}/reset"
GRADE_ENDPOINT = f"{API_ROOT}/grade-response"

st.set_page_config(page_title="DSA Planning Chatbot", page_icon="🧭")
st.title("DSA Planning Chatbot")
st.caption("Quick demo UI backed by a Flask API and LangGraph agent.")

if "messages" not in st.session_state:
    st.session_state["messages"] = []
if "developer_payload" not in st.session_state:
    st.session_state["developer_payload"] = []
if "ablation_prompt_template" not in st.session_state:
    st.session_state["ablation_prompt_template"] = DEFAULT_SYSTEM_PROMPT_TEMPLATE
if "ablation_reasoning_enabled" not in st.session_state:
    st.session_state["ablation_reasoning_enabled"] = False
if "ablation_retriever_enabled" not in st.session_state:
    st.session_state["ablation_retriever_enabled"] = True
if "ground_truth_text" not in st.session_state:
    st.session_state["ground_truth_text"] = ""
if "evaluation_results" not in st.session_state:
    st.session_state["evaluation_results"] = []
if "evaluation_running" not in st.session_state:
    st.session_state["evaluation_running"] = False
if "evaluation_timeouts" not in st.session_state:
    st.session_state["evaluation_timeouts"] = []
if "test_grader_result" not in st.session_state:
    st.session_state["test_grader_result"] = None
if "test_grader_error" not in st.session_state:
    st.session_state["test_grader_error"] = ""

st.markdown(
    """
    <style>
    .assistant-loading {
        display: inline-flex;
        gap: 0.35rem;
        align-items: center;
        padding: 0.1rem 0;
    }
    .assistant-loading .dot {
        width: 0.35rem;
        height: 0.35rem;
        background-color: currentColor;
        border-radius: 50%;
        opacity: 0.25;
        animation: assistant-loading-bounce 1.4s infinite ease-in-out;
    }
    .assistant-loading .dot:nth-child(2) {
        animation-delay: 0.2s;
    }
    .assistant-loading .dot:nth-child(3) {
        animation-delay: 0.4s;
    }
    @keyframes assistant-loading-bounce {
        0%, 80%, 100% {
            opacity: 0.25;
            transform: scale(0.75);
        }
        40% {
            opacity: 1;
            transform: scale(1);
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

ASSISTANT_LOADING_HTML = (
    "<div class=\"assistant-loading\">"
    "<span class=\"dot\"></span>"
    "<span class=\"dot\"></span>"
    "<span class=\"dot\"></span>"
    "</div>"
)

EVALUATION_QUESTIONS = [
    {
        "id": 1,
        "question": "List all the Level 2000 core modules for the DSA major.",
        "ground_truth": (
            "Must include: CS2040, DSA2101, DSA2102, MA2001, MA2002, MA2311, ST2131, "
            "ST2132; Total Units = 32 MCs."
        ),
    },
    {
        "id": 2,
        "question": "What are the Level 3000 modules I need to take for DSA?",
        "ground_truth": "Must include: CS3244, DSA3101, DSA3102, ST3131; Total Units = 16 MCs.",
    },
    {
        "id": 3,
        "question": "Do I have to do a specialisation in DSA to graduate?",
        "ground_truth": (
            "No, specialisations are optional; Answer should mention Operations Research and Statistical "
            "Methodology as the optional specialisations."
        ),
    },
    {
        "id": 4,
        "question": "How many MCs do I need to graduate from the DSA programme?",
        "ground_truth": (
            "60 units for major requirements, within the total CHS framework (~160 MCs including CHS). "
            "Must include Level 1000–4000 breakdown."
        ),
    },
    {
        "id": 5,
        "question": "Explain the difference between Option A and Option B in Level 4000 requirements.",
        "ground_truth": (
            "Must mention Option A, Option B, Honours Project; Explain that Option A = 2 modules and "
            "Option B = Honours Project (DSA4288 variant)."
        ),
    },
    {
        "id": 6,
        "question": "What modules are in the Operations Research specialisation?",
        "ground_truth": (
            "Must include: MA3252, MA3238, MA4260, MA4270, DSA4288M; Total Units = 20 MCs."
        ),
    },
    {
        "id": 7,
        "question": "List the Statistical Methodology specialisation modules.",
        "ground_truth": (
            "Must include: ST3232, ST3248, ST4231, ST4234, ST4253, DSA4288S"
        ),
    },
    {
        "id": 8,
        "question": "What are the CHS Common Curriculum pillars?",
        "ground_truth": (
            "Must include: Data Literacy, Communities and Engagement, Artificial Intelligence, Design "
            "Thinking, Digital Literacy, Writing."
        ),
    },
    {
        "id": 9,
        "question": "Give examples of Communities and Engagement courses I can take.",
        "ground_truth": (
            "Must minimally include: GEN2061X, GEN2061Y, CLC1101, CLC2204; Note that they are "
            "service-learning or project-based."
        ),
    },
    {
        "id": 10,
        "question": "Which CHS pillar does DSA1101 fulfil?",
        "ground_truth": "It fulfils the Data Literacy pillar.",
    },
    {
        "id": 11,
        "question": "Which semester should I take Communities and Engagement modules?",
        "ground_truth": (
            "Semester 1: Modules such as GEN2050X (Teach SG), GEN2060X (Reconnect SeniorsSG), "
            "GEN2061X (Support Healthy AgeingSG), GEN2062X (Community Activities for Seniors with SG "
            "Cares), and GEN2070X (Community Link (Comlink) Befrienders) are offered in Semester 1.\n"
            "Semester 2: Modules such as GEN2050Y (Teach SG), GEN2060Y (Reconnect SeniorsSG), GEN2061Y "
            "(Support Healthy AgeingSG), GEN2062Y (Community Activities for Seniors with SG Cares), and "
            "GEN2070Y (Community Link (Comlink) Befrienders) are offered in Semester 2."
        ),
    },
    {
        "id": 12,
        "question": "When can I take the Interdisciplinary CHS courses?",
        "ground_truth": "Recommended Years 3 and 4",
    },
    {
        "id": 13,
        "question": (
            "I’m a year 2 DSA student who has taken DSA1101 and MA2001 — what modules do you suggest I take "
            "next in Semester 1 to fulfil a 20 MC workload?"
        ),
        "ground_truth": (
            "Must suggest 5 from this list: CS2040, DSA2101, DSA2102, MA2311, ST2131, ST2132 or MA2002; "
            "Total ≈ 20 MCs."
        ),
    },
    {
        "id": 14,
        "question": "What modules should I take next semester?",
        "ground_truth": (
            "Should ask clarifying questions about AY, completed modules, and interests instead of guessing."
        ),
    },
    {
        "id": 15,
        "question": "Book a Grab ride to NUS for me.",
        "ground_truth": "Politely refuse and redirect to academic planning topics.",
    },
    {
        "id": 16,
        "question": "What are the prerequisites for XYZ9999?",
        "ground_truth": (
            "Apologise and explain module not found; suggest verifying code or AY; do not hallucinate."
        ),
    },
    {
        "id": 17,
        "question": "What is the timetable for DSA4213?",
        "ground_truth": "Should default to the current AY 25/26",
    },
    {
        "id": 18,
        "question": "What’s the weather in Singapore like?",
        "ground_truth": "Politely refuse and redirect to academic planning topics.",
    },
    {
        "id": 19,
        "question": "What are the prerequisites for DSA1111 Data Science for Beginners?",
        "ground_truth": (
            "Apologise and explain module not found; suggest verifying code or AY; do not hallucinate."
        ),
    },
    {
        "id": 20,
        "question": "What are the prerequisites for DSA4213?",
        "ground_truth": (
            "If undertaking an Undergraduate Degree THEN( must have completed MA2001 at a grade of at least D AND "
            "must have completed 1 of MA2104/MA2311 at a grade of at least D AND must have completed 1 of ST2137 at a "
            "grade of at least D, any Courses beginning with CS2040 at a grade of at least D)"
        ),
    },
    {
        "id": 21,
        "question": "What is the timetable for DSA4213 like in Semester 1?",
        "ground_truth": (
            "Lecture 1: Every week from Week 1 to Week 13, on Fridays from 2:00 PM to 4:00 PM in UT-AUD2.\n"
            "Lecture 1: Every week from Week 1 to Week 13, on Tuesdays from 5:00 PM to 7:00 PM in UT-AUD1."
        ),
    },
]


def _current_ablation_sheet_name(*, reasoning_enabled: bool, retriever_enabled: bool) -> str:
    """Return the worksheet name for the active ablation configuration."""

    reasoning_flag = "reasoning_on" if reasoning_enabled else "reasoning_off"
    retriever_flag = "retriever_on" if retriever_enabled else "retriever_off"
    return f"{reasoning_flag}_{retriever_flag}"


def _serialise_reasoning_trace(value: object) -> str:
    """Convert the grader reasoning trace into a spreadsheet-friendly string."""

    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except TypeError:
            return str(value)
    return str(value)


def _evaluation_workbook_path() -> Path:
    """Return the path to the shared evaluation workbook."""

    return Path(__file__).resolve().parent / "evaluation_results.xlsx"


def _write_evaluation_csv(
    results: list[dict],
    *,
    reasoning_enabled: bool,
    retriever_enabled: bool,
) -> tuple[Path, str] | None:
    """Export evaluation outcomes to a shared workbook with per-ablation sheets."""

    if not results:
        return None

    workbook_path = _evaluation_workbook_path()
    sheet_name = _current_ablation_sheet_name(
        reasoning_enabled=reasoning_enabled,
        retriever_enabled=retriever_enabled,
    )

    fieldnames = [
        "question",
        "assistant_response",
        "response_time_seconds",
        "ground_truth",
        "grader_reasoning_trace",
        "grader_accuracy",
        "grader_relevance",
        "grader_coherence",
        "final_score",
    ]

    rows: list[dict[str, object]] = []
    for item in results:
        evaluation = item.get("evaluation") or {}
        scores = evaluation.get("scores") or {}
        rows.append(
            {
                "question": item.get("question", ""),
                "assistant_response": item.get("answer", ""),
                "response_time_seconds": item.get("response_time"),
                "ground_truth": item.get("ground_truth", ""),
                "grader_reasoning_trace": _serialise_reasoning_trace(
                    evaluation.get("grader_reasoning")
                ),
                "grader_accuracy": scores.get("accuracy"),
                "grader_relevance": scores.get("relevance"),
                "grader_coherence": scores.get("coherence"),
                "final_score": evaluation.get("total"),
            }
        )

    try:
        workbook = load_workbook(workbook_path)
    except FileNotFoundError:
        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        workbook.remove(worksheet)

    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(fieldnames)
    for row in rows:
        worksheet.append([row.get(column) for column in fieldnames])

    workbook.save(workbook_path)

    return workbook_path, sheet_name


def _merge_history(existing: list[dict[str, str]], new: list[dict[str, str]]) -> list[dict[str, str]]:
    """Combine truncated API history with the full UI history."""

    if not existing:
        return list(new)

    max_overlap = min(len(existing), len(new))
    overlap = 0
    for size in range(max_overlap, 0, -1):
        if existing[-size:] == new[:size]:
            overlap = size
            break

    if overlap == 0 and new:
        # If no overlap is detected we assume the API returned a new session.
        return list(new)

    return existing + new[overlap:]


def _render_evaluation(evaluation: dict) -> None:
    """Display the grader scores inside the active chat container."""

    if not evaluation:
        return

    error_message = evaluation.get("error")
    if error_message:
        st.warning(f"Grader could not produce a score: {error_message}")
        return

    scores = evaluation.get("scores", {})
    accuracy = scores.get("accuracy")
    relevance = scores.get("relevance")
    coherence = scores.get("coherence")
    total = evaluation.get("total")

    def _format_score(value: float | None) -> str:
        if value is None:
            return "–"
        return f"{value:.1f}"

    lines = [
        "**Grader scores**",
        f"- Accuracy: {_format_score(accuracy)}",
        f"- Relevance: {_format_score(relevance)}",
        f"- Fluency & coherence: {_format_score(coherence)}",
    ]
    if total is not None:
        lines.extend(["", f"**Final score:** {total:.1f} / 3"])

    st.markdown("\n".join(lines))


ASSISTANT_AVATAR = "🤖"
AVATARS = {
    "assistant": ASSISTANT_AVATAR,
    "user": "🧑‍💻",
}


def _render_developer_payload(payload: dict) -> None:
    """Display developer view metadata inside an expander stack."""

    with st.expander("Model input", expanded=False):
        st.json(payload.get("model_input", []))

    stream_events = payload.get("stream_events", [])
    if stream_events:
        st.markdown("Stream trace")
        for event_index, event in enumerate(stream_events, start=1):
            label = event.get("type", f"event {event_index}")
            with st.expander(f"Event {event_index}: {label}", expanded=False):
                st.json(event)

    router_decision = payload.get("router_decision")
    retrieved_docs = payload.get("retrieved_docs")
    if router_decision or retrieved_docs:
        st.markdown("Retrieval summary")
        if router_decision:
            st.write(f"Router decision: **{router_decision}**")
        if retrieved_docs:
            with st.expander("Retrieved documents", expanded=False):
                st.json(retrieved_docs)

    grader_info = payload.get("grader")
    if grader_info:
        st.markdown("Grader details")
        prompt_messages = grader_info.get("messages")
        if prompt_messages:
            with st.expander("Grader prompt", expanded=False):
                st.json(prompt_messages)
        parsed_scores = grader_info.get("parsed_scores")
        if parsed_scores:
            st.json(parsed_scores)
        reasoning_traces = grader_info.get("reasoning_traces")
        if reasoning_traces:
            with st.expander("Reasoning traces", expanded=True):
                if isinstance(reasoning_traces, (dict, list)):
                    st.json(reasoning_traces)
                else:
                    st.code(str(reasoning_traces))
        additional_kwargs = grader_info.get("additional_kwargs")
        if additional_kwargs:
            with st.expander("Additional kwargs", expanded=False):
                st.json(additional_kwargs)
        response_metadata = grader_info.get("response_metadata")
        if response_metadata:
            with st.expander("Response metadata", expanded=False):
                st.json(response_metadata)
        raw_response = grader_info.get("raw_response")
        if raw_response:
            st.code(raw_response, language="json")

    with st.expander("Stored chat state", expanded=False):
        st.json(payload.get("stored_state", []))

    configuration = payload.get("configuration")
    if configuration:
        with st.expander("Ablation configuration", expanded=False):
            st.json(configuration)


def _render_evaluation_result(result: dict, *, index: int, developer_view: bool) -> None:
    """Pretty-print a single evaluation outcome."""

    question = result.get("question", "")
    st.markdown(f"### Question {index}: {question}")

    error_message = result.get("error")
    if error_message:
        st.error(error_message)
        return

    assistant_answer = result.get("answer", "")
    st.markdown("**Assistant response**")
    if assistant_answer:
        st.markdown(assistant_answer)
    else:
        st.info("No response was returned by the assistant.")

    response_time_value = result.get("response_time")
    if response_time_value is not None:
        st.caption(f"Response time: {response_time_value:.2f}s")

    ground_truth = result.get("ground_truth")
    if ground_truth:
        with st.expander("Ground truth reference", expanded=False):
            st.write(ground_truth)

    evaluation_meta = result.get("evaluation") or {}
    if evaluation_meta:
        _render_evaluation(evaluation_meta)

        grader_prompt = evaluation_meta.get("grader_prompt")
        if grader_prompt:
            with st.expander("Grader prompt", expanded=False):
                st.code(grader_prompt)

        reasoning_traces = evaluation_meta.get("grader_reasoning")
        if reasoning_traces is not None:
            with st.expander("Grader reasoning trace", expanded=True):
                if isinstance(reasoning_traces, (dict, list)):
                    st.json(reasoning_traces)
                else:
                    st.code(str(reasoning_traces))
    else:
        st.warning("No grading information was returned for this question.")

    if developer_view:
        developer_payload = result.get("developer")
        if developer_payload:
            st.divider()
            st.subheader("Developer details")
            _render_developer_payload(developer_payload)


def _render_evaluation_results(results: list[dict], *, developer_view: bool) -> None:
    """Render the complete evaluation report in order."""

    if not results:
        st.info("Run the evaluation suite to see grading outcomes.")
        return

    for display_index, item in enumerate(results, start=1):
        _render_evaluation_result(item, index=display_index, developer_view=developer_view)
        if display_index != len(results):
            st.divider()


with st.sidebar:
    st.header("Session Controls")
    developer_view_enabled = st.toggle("Developer view", value=False, help="Show request and tool trace details.")
    if st.button("Reset conversation"):
        try:
            requests.post(RESET_ENDPOINT, timeout=10)
            st.session_state["messages"] = []
            st.session_state["developer_payload"] = []
            st.success("Conversation cleared.")
            st.rerun()
        except requests.RequestException as exc:
            st.error(f"Failed to reset chat: {exc}")

    st.divider()
    st.header("Ablation controls")
    st.caption("Adjust the assistant configuration for quick ablation studies.")
    st.text_area(
        "Assistant system prompt",
        key="ablation_prompt_template",
        height=220,
    )
    st.toggle(
        "Enable Qwen reasoning",
        key="ablation_reasoning_enabled",
        value=st.session_state["ablation_reasoning_enabled"],
        help="Toggle the model's built-in reasoning capabilities.",
    )
    st.toggle(
        "Enable retriever tool",
        key="ablation_retriever_enabled",
        value=st.session_state["ablation_retriever_enabled"],
        help="Allow the assistant to call the curriculum retriever node.",
    )
    st.divider()
    st.header("Grading")
    st.caption(
        "Provide a reference answer if you would like the assistant's next response to be graded."
    )
    st.text_area(
        "Ground truth reference (optional)",
        key="ground_truth_text",
        height=180,
        help="Leave blank to skip grading.",
    )

chat_tab, evaluation_tab, test_grader_tab = st.tabs(["Chat", "Evaluation", "Test grader"])

with chat_tab:
    for item in st.session_state.get("messages", []):
        role = item.get("role", "assistant")
        content = item.get("content", "")
        avatar = AVATARS.get(role, "💬")
        chat_role = role if role in ("user", "assistant") else "assistant"
        with st.chat_message(chat_role, avatar=avatar):
            st.markdown(content)
            metadata = item.get("metadata") or {}
            evaluation_meta = metadata.get("evaluation") if isinstance(metadata, dict) else None
            if evaluation_meta:
                _render_evaluation(evaluation_meta)
            if role == "assistant" and isinstance(metadata, dict):
                response_time_value = metadata.get("response_time")
                if response_time_value is not None:
                    st.caption(f"Response time: {response_time_value:.2f}s")

    user_prompt = st.chat_input(
        placeholder="e.g. What is the timetable for DSA4213 in semester 1?",
    )

    if user_prompt and user_prompt.strip():
        user_prompt = user_prompt.strip()
        st.session_state["messages"].append({"role": "user", "content": user_prompt})
        with st.chat_message("user", avatar=AVATARS["user"]):
            st.markdown(user_prompt)

        with st.chat_message("assistant", avatar=ASSISTANT_AVATAR):
            message_placeholder = st.empty()
            message_placeholder.markdown(ASSISTANT_LOADING_HTML, unsafe_allow_html=True)
            try:
                response = requests.post(
                    CHAT_ENDPOINT,
                    json={
                        "prompt": user_prompt,
                        "developer_view": developer_view_enabled,
                        "system_prompt_template": st.session_state["ablation_prompt_template"],
                        "enable_reasoning": st.session_state["ablation_reasoning_enabled"],
                        "enable_retriever": st.session_state["ablation_retriever_enabled"],
                        "ground_truth": st.session_state.get("ground_truth_text", ""),
                    },
                    timeout=(10, 600),
                )
                response.raise_for_status()
                data = response.json()
                api_history = data.get("history", [])
                st.session_state["messages"] = _merge_history(
                    st.session_state.get("messages", []), api_history
                )
                developer_info = data.get("developer_view")
                if developer_info:
                    st.session_state["developer_payload"].append(developer_info)

                assistant_reply = ""
                last_assistant_message: dict[str, object] | None = None
                for message in reversed(st.session_state["messages"]):
                    if message.get("role") == "assistant":
                        assistant_reply = message.get("content", "")
                        last_assistant_message = message
                        break
                if not assistant_reply:
                    assistant_reply = data.get("answer", "")
                if assistant_reply:
                    message_placeholder.markdown(assistant_reply)
                else:
                    message_placeholder.empty()
                response_time_value = data.get("response_time")
                if last_assistant_message and response_time_value is not None:
                    metadata = last_assistant_message.setdefault("metadata", {})
                    if isinstance(metadata, dict):
                        metadata["response_time"] = response_time_value
                if response_time_value is not None:
                    st.caption(f"Response time: {response_time_value:.2f}s")
                if last_assistant_message:
                    metadata = last_assistant_message.get("metadata") or {}
                    evaluation_meta = metadata.get("evaluation") if isinstance(metadata, dict) else None
                    if evaluation_meta:
                        _render_evaluation(evaluation_meta)
            except requests.RequestException as exc:
                error_message = f"Failed to contact the chatbot API: {exc}"
                message_placeholder.markdown(error_message)
                st.session_state["messages"].append({"role": "assistant", "content": error_message})

    if developer_view_enabled and st.session_state.get("developer_payload"):
        st.divider()
        st.subheader("Developer details")
        for index, payload in enumerate(st.session_state["developer_payload"], start=1):
            st.markdown(f"**Interaction {index}**")
            _render_developer_payload(payload)

with evaluation_tab:
    st.subheader("Evaluation harness")
    st.caption(
        "Runs the predefined evaluation questions using the current ablation settings and displays grading details."
    )

    run_evaluation = st.button(
        "Run evaluation suite",
        disabled=st.session_state["evaluation_running"],
    )
    if run_evaluation and not st.session_state["evaluation_running"]:
        st.session_state["evaluation_results"] = []
        st.session_state["evaluation_timeouts"] = []
        st.session_state["evaluation_running"] = True

    progress_placeholder = st.empty()
    status_placeholder = st.empty()
    results_placeholder = st.empty()

    if st.session_state["evaluation_running"]:
        total_questions = len(EVALUATION_QUESTIONS)
        progress_bar = progress_placeholder.progress(0)
        try:
            for index, item in enumerate(EVALUATION_QUESTIONS, start=1):
                question = item.get("question", "")
                ground_truth = item.get("ground_truth", "")
                status_placeholder.info(
                    f"Evaluating question {index}/{total_questions}: {question}"
                )

                result_entry: dict[str, object] = {
                    "id": item.get("id"),
                    "question": question,
                    "ground_truth": ground_truth,
                }

                try:
                    requests.post(RESET_ENDPOINT, timeout=10)
                except requests.RequestException as exc:
                    result_entry["error"] = (
                        f"Failed to reset the conversation before evaluation: {exc}"
                    )
                else:
                    try:
                        response = requests.post(
                            CHAT_ENDPOINT,
                            json={
                                "prompt": question,
                                "developer_view": developer_view_enabled,
                                "system_prompt_template": st.session_state["ablation_prompt_template"],
                                "enable_reasoning": st.session_state["ablation_reasoning_enabled"],
                                "enable_retriever": st.session_state["ablation_retriever_enabled"],
                                "ground_truth": ground_truth,
                            },
                            timeout=(10, 600),
                        )
                        response.raise_for_status()
                        data = response.json()
                        result_entry["answer"] = data.get("answer", "")
                        result_entry["response_time"] = data.get("response_time")
                        result_entry["evaluation"] = data.get("evaluation") or {}
                        developer_info = data.get("developer_view")
                        if developer_info:
                            result_entry["developer"] = developer_info
                    except requests.Timeout:
                        timeout_message = (
                            f"API request timed out for question {item.get('id')}: {question}"
                        )
                        result_entry["error"] = timeout_message
                        st.session_state["evaluation_timeouts"].append(timeout_message)
                    except requests.RequestException as exc:
                        result_entry["error"] = f"Failed to evaluate question: {exc}"
                    except ValueError as exc:
                        result_entry["error"] = f"Invalid JSON response: {exc}"

                st.session_state["evaluation_results"].append(result_entry)

                progress_bar.progress(index / total_questions)
                results_placeholder.empty()
                with results_placeholder.container():
                    _render_evaluation_results(
                        st.session_state["evaluation_results"],
                        developer_view=developer_view_enabled,
                    )

            try:
                workbook_info = _write_evaluation_csv(
                    st.session_state["evaluation_results"],
                    reasoning_enabled=st.session_state["ablation_reasoning_enabled"],
                    retriever_enabled=st.session_state["ablation_retriever_enabled"],
                )
            except OSError as exc:
                status_placeholder.warning(
                    f"Evaluation suite completed but could not write workbook: {exc}"
                )
            else:
                if workbook_info is not None:
                    workbook_path, sheet_name = workbook_info
                    status_placeholder.success(
                        "Evaluation suite completed. Results saved to "
                        f"{workbook_path.name} (sheet: {sheet_name})."
                    )
                else:
                    status_placeholder.success("Evaluation suite completed.")
        finally:
            st.session_state["evaluation_running"] = False
            progress_bar.progress(1.0)
    else:
        with results_placeholder.container():
            _render_evaluation_results(
                st.session_state.get("evaluation_results", []),
                developer_view=developer_view_enabled,
            )

    if st.session_state.get("evaluation_timeouts"):
        timeout_list = "\n".join(
            f"- {message}" for message in st.session_state["evaluation_timeouts"]
        )
        st.warning(
            "One or more questions exceeded the API timeout:\n" + timeout_list
        )

with test_grader_tab:
    st.subheader("Test grader")
    st.caption(
        "Manually provide a question, ground truth answer, and assistant response to run the grading pipeline."
    )

    question_input = st.text_area(
        "Question",
        key="test_grader_question",
        placeholder="Enter the question you want the assistant graded against.",
    )
    ground_truth_input = st.text_area(
        "Ground truth",
        key="test_grader_ground_truth",
        placeholder="Provide the authoritative answer or rubric for grading.",
        height=150,
    )
    assistant_response_input = st.text_area(
        "Assistant response",
        key="test_grader_answer",
        placeholder="Paste the assistant's response to be graded.",
        height=200,
    )
    developer_view_requested = st.checkbox(
        "Show developer details for the grader",
        key="test_grader_developer_view",
    )

    if st.button("Grade response", key="grade_response_button"):
        if not question_input.strip():
            st.session_state["test_grader_error"] = "Question is required for grading."
            st.session_state["test_grader_result"] = None
        elif not ground_truth_input.strip():
            st.session_state["test_grader_error"] = "Ground truth is required for grading."
            st.session_state["test_grader_result"] = None
        elif not assistant_response_input.strip():
            st.session_state["test_grader_error"] = "Assistant response is required for grading."
            st.session_state["test_grader_result"] = None
        else:
            try:
                response = requests.post(
                    GRADE_ENDPOINT,
                    json={
                        "question": question_input,
                        "ground_truth": ground_truth_input,
                        "answer": assistant_response_input,
                        "developer_view": developer_view_requested,
                    },
                    timeout=120,
                )
                response.raise_for_status()
            except requests.RequestException as exc:
                st.session_state["test_grader_error"] = (
                    f"Failed to grade the response: {exc}"
                )
                st.session_state["test_grader_result"] = None
            else:
                try:
                    data = response.json()
                except ValueError as exc:
                    st.session_state["test_grader_error"] = (
                        f"Invalid JSON response from grader: {exc}"
                    )
                    st.session_state["test_grader_result"] = None
                else:
                    api_error = data.get("error")
                    if api_error:
                        st.session_state["test_grader_error"] = str(api_error)
                        st.session_state["test_grader_result"] = None
                    else:
                        st.session_state["test_grader_error"] = ""
                        st.session_state["test_grader_result"] = {
                            "evaluation": data.get("evaluation") or {},
                            "developer": data.get("developer_view"),
                            "developer_view_enabled": developer_view_requested,
                        }

    if st.session_state.get("test_grader_error"):
        st.error(st.session_state["test_grader_error"])

    test_grader_result = st.session_state.get("test_grader_result")
    if test_grader_result:
        evaluation_payload = test_grader_result.get("evaluation") or {}
        if evaluation_payload:
            _render_evaluation(evaluation_payload)

            reasoning_traces = evaluation_payload.get("grader_reasoning")
            if reasoning_traces is not None:
                with st.expander("Grader reasoning trace", expanded=True):
                    if isinstance(reasoning_traces, (dict, list)):
                        st.json(reasoning_traces)
                    else:
                        st.code(str(reasoning_traces))
        else:
            st.info("The grader did not return a score for this response.")

        if (
            test_grader_result.get("developer_view_enabled")
            and test_grader_result.get("developer")
        ):
            st.divider()
            st.subheader("Grader developer details")
            _render_developer_payload(test_grader_result["developer"])
